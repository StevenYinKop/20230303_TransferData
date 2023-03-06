import os
import sys
# 解析xlsx所用的库
from openpyxl import load_workbook, Workbook
# 解析xls所用的库
from xlrd import open_workbook

# 定义标题名称
title_author = '数据采集人'
title_file_path = '数据文件名称'
title_sheet_name = 'Sheet页名称(默认为第一页）'
title_indicator_name = '采集指标名称'
title_year = '数据年度'
title_start_at = '采集起始行/列'
title_end_at = '采集终止行/列'
title_field_index = '字段名称列/行号'
title_data_index = '数据列/行号'
title_is_horizontal_or_vertical = '横表/竖表'

# 解析当前python脚本所在的路径
script_path = os.path.abspath(__file__)
# 默认的数据路径，设定为：/python脚本所在路径/工作目录/
base_dir = f'{os.path.dirname(script_path)}{os.sep}工作目录'

# 定义输出的文件名
output_filename = f'{base_dir}{os.sep}total.xlsx'
# 定义错误日志的文件名
error_filename = f'{base_dir}{os.sep}错误日志.txt'
# 定义收集错误日志所用的列表
error_list = []

# 定义索引文件的标题栏，后续用来方便读取索引文件的数据
title_list = [
    title_author,
    title_file_path,
    title_sheet_name,
    title_indicator_name,
    title_year,
    title_start_at,
    title_end_at,
    title_field_index,
    title_data_index,
    title_is_horizontal_or_vertical
]

"""声明一个结果集，后续的结果都会存在这个列表中"""
data_list = [
    [
        '数据采集人',
        '数据指标名称',
        '字段名称',
        '数据年度',
        '数据']
]

"""
格式化错误日志，使得错误日志中，每一行都可以有统一的前缀
类似于：[-----第5行-----]：文件解析异常！
"""
def get_format_error_cell(error_str, row_number):
    return f'[-----第{row_number}行-----]：{error_str}'


"""
将错误日志输出到文件中，输出完毕后，终止当前程序
"""
def output_error_list():
    # 如果没有错误日志，那么不执行后续的操作了。
    if len(error_list) == 0:
        return
    error_file = open(error_filename, 'w')
    # 逐行将错误内容写入文件
    for error_item in error_list:
        error_file.write(f'{error_item}\n')
    print(f"程序执行过程中发现错误，请查看{error_filename}")
    error_file.close()
    # 退出python程序
    sys.exit()
    # exit()

"""
格式化索引: 如果传入的"采集起始行/列"和"采集终止行/列"超过了excel中已有的最大长度，则以数据最大长度为准：

如："采集终止行/列"为100，但是数据最多只有50条，那么强制将"采集终止行/列"读取为50
"""
def standardize_index(index_row_dict, get_max):
    return max(int(index_row_dict[title_start_at]), 1), \
           min(int(index_row_dict[title_end_at]), get_max())

"""
组装最终输入到结果文件的数据
"""
def assemble_data_list_item(field, data, index_dict):
    return [
        index_dict[title_author], # 数据采集人
        index_dict[title_indicator_name],  # 数据指标名称
        str(field),  # 字段名称
        int(index_dict[title_year]),  # 数据年度
        data  # 数据
    ]


"""
校验索引文件的格式, 防止索引文件中，缺少一些必要的字段，导致程序无法正确执行
"""
def validate_index_file_parameters(headers, index_filename):
    print(f'正在检查索引文件"{index_filename}"中的数据有效性...')
    for title_str in title_list:
        if title_str not in headers:
            error_list.append(f'索引文件"{index_filename}"中缺失必要参数列："{title_str}"，请检查。')
    output_error_list()


"""
验证索引文件中的每一个字段是否能够正常读取
"""
def validate_index_data(index_dict, row):
    error_sub_list = []
    for title_str in [title_year, title_start_at, title_end_at, title_data_index, title_field_index]:
        try:
            # 如果年份，起止行等数据，不是数值格式(1,2,3)，而是一些奇怪的值(任意的文字或者字母)，则将错误信息放入错误列表中
            int(float((index_dict[title_str])))
        except Exception as e:
            error_sub_list.append(get_format_error_cell(f' "{title_str}" 不是一个数字格式，无法解析！', row))
    if index_dict[title_is_horizontal_or_vertical] not in ['横表', '竖表']:
        error_sub_list.append(get_format_error_cell(f' "{title_is_horizontal_or_vertical}"中只能填"横表"或者"竖表"！', row))
        # 如果数据文件不存在，则将错误信息存入错误列表中
    if not os.path.exists(f'{base_dir}{os.sep}{str(index_dict[title_file_path])}'):
        error_sub_list.append(get_format_error_cell(f' "{title_file_path}" 文件不存在，无法解析！', row))
    else:
        # 检查数据文件中，是否存在相应的Sheet页。
        file_path = f'{base_dir}{os.sep}{str(index_dict[title_file_path])}'
        _, extension = os.path.splitext(file_path)
        if extension == '.xlsx':
            workbook = load_workbook(file_path)
            if index_dict[title_sheet_name] not in workbook.sheetnames and index_dict[title_sheet_name] is not None:
                error_sub_list.append(
                    get_format_error_cell(f'{title_file_path}文件中不存在名为{index_dict[title_sheet_name]}的Sheet页，无法解析！', row))
            workbook.close()
        elif extension == '.xls':
            workbook = open_workbook(file_path)
            if index_dict[title_sheet_name] not in workbook.sheet_names() and str(index_dict[title_sheet_name]) not in [
                'None', '']:
                error_sub_list.append(
                    get_format_error_cell(f'{title_file_path}文件中不存在名为{index_dict[title_sheet_name]}的Sheet页，无法解析！', row))
            workbook.release_resources()
    # 如果错误列表中存在数据，则说明这一条数据是无法正常解析的，将会忽略对该行的操作
    if len(error_sub_list) != 0:
        error_sub_list.append(get_format_error_cell(f'跳过第{row}行...\n', row))
    error_list.extend(error_sub_list)
    return len(error_sub_list) == 0


"""
定义一个解析数据文件的函数
"""
def parse_data_file(file_path, index_row_dict):
    sub_data_list = []
    print(f'开始处理文件{file_path}...')
    _, extension = os.path.splitext(file_path)
    field_index = int(index_row_dict[title_field_index])
    data_index = int(index_row_dict[title_data_index])
    # 根据不同的文件名后缀，选择不同的库对excel进行操作
    if extension == '.xlsx':
        print('文件是xlsx格式，使用新excel的解析方式...请稍后...')
        workbook = load_workbook(file_path)
        # 如果索引文件中指定了Sheet名，那么用他指定的名字，否则选择excel中的第一个sheet页
        sheet = workbook[index_row_dict[title_sheet_name]] \
            if index_row_dict[title_sheet_name] not in [None, ''] \
            else workbook.worksheets[0]
        print(f'文件加载成功，当前操作的sheet页为{sheet.title}')
        if index_row_dict[title_is_horizontal_or_vertical] == '横表':
            # 横表操作
            start_at, end_at = standardize_index(index_row_dict, lambda: sheet.max_column)
            print(f'以"{index_row_dict[title_is_horizontal_or_vertical]}"'
                  f'的方式解析当前sheet页的第{start_at}列至第{end_at}列...')
            # 因为是横表，所以逐列对数据进行读取
            for col in range(start_at, end_at + 1):
                print(f'开始解析第{col}列...')
                item = assemble_data_list_item(sheet.cell(field_index, col).value,
                                               sheet.cell(data_index, col).value,
                                               index_row_dict)
                sub_data_list.append(item)
                print(f'第{col}列解析完毕！, 数据为：{item}')
        elif index_row_dict[title_is_horizontal_or_vertical] == '竖表':
            # 竖表操作
            start_at, end_at = standardize_index(index_row_dict, lambda: sheet.max_row)
            print(f'以"{index_row_dict[title_is_horizontal_or_vertical]}"'
                  f'的方式解析当前sheet页的第{start_at}行至第{end_at}行...')
            for row in range(start_at, end_at + 1):  # sheet.iter_rows(min_col=start_at, max_col=end_at):
                print(f'开始解析第{row}行...')
                # sheet[]
                item = assemble_data_list_item(sheet.cell(row, field_index).value,
                                               sheet.cell(row, data_index).value,
                                               index_row_dict)
                sub_data_list.append(item)
                print(f'第{row}行解析完毕！, 数据为：{item}')

    elif extension == '.xls':
        print('文件是xls格式，使用旧excel的解析方式...请稍后...')
        workbook = open_workbook(file_path)
        sheet = workbook.sheet_by_name(index_row_dict[title_sheet_name]) \
            if index_row_dict[title_sheet_name] not in [None, ''] \
            else workbook.sheet_by_index(0)
        print(f'文件加载成功，当前操作的sheet页为{sheet.name}')

        if index_row_dict[title_is_horizontal_or_vertical] == '横表':
            start_at, end_at = standardize_index(index_row_dict, lambda: sheet.ncols)
            print(f'以"{index_row_dict[title_is_horizontal_or_vertical]}"'
                  f'的方式解析当前sheet页的第{start_at}列至第{end_at}列...')
            for col in range(start_at, end_at + 1):
                print(f'开始解析第{col}列...')
                item = assemble_data_list_item(sheet.cell_value(field_index - 1, col - 1),
                                               sheet.cell_value(data_index - 1, col - 1),
                                               index_row_dict)
                sub_data_list.append(item)
                print(f'第{col}列解析完毕！, 数据为：{item}')
        elif index_row_dict[title_is_horizontal_or_vertical] == '竖表':
            start_at, end_at = standardize_index(index_row_dict, lambda: sheet.nrows)
            print(f'以"{index_row_dict[title_is_horizontal_or_vertical]}"'
                  f'的方式解析当前sheet页的第{start_at}行至第{end_at}行...')
            for row in range(start_at, end_at + 1):
                print(f'开始解析第{row}行...')
                item = assemble_data_list_item(sheet.cell_value(row - 1, field_index - 1),
                                               sheet.cell_value(row - 1, data_index - 1),
                                               index_row_dict)
                sub_data_list.append(item)
                print(f'第{row}行解析完毕！, 数据为：{item}')
    else:
        print(f'\n错误！！！ === 输入的文件" {file_path} "不是一个标准的excel，请使用.xls或者.xlsx文件 === ', end='\n\n')
    return sub_data_list


# 定义一个解析索引文件的函数
def parse_index_file(index_filename):
    _, extension = os.path.splitext(index_filename)
    # 如果是xlsx后缀名的文件，走这个if逻辑
    if extension == '.xlsx':
        workbook = load_workbook(index_filename)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        validate_index_file_parameters(headers, index_filename)
        for row in sheet.iter_rows(min_row=2):
            # 解析索引文件，将索引文件的每一行作为分别保存下来，存入index_row_dict
            index_row_dict = {headers[i]: row[i].value for i in range(len(headers))}
            if validate_index_data(index_row_dict, row[0].row):
                # 代码能够走到这里，说明索引文件的数据是可以正常被解析到的，那么开始解析数据文件中的数据
                data_list.extend(
                    parse_data_file(f'{base_dir}{os.sep}{index_row_dict[title_file_path]}', index_row_dict))

    # 如果是xls后缀名的文件，走以下逻辑
    elif extension == '.xls':
        workbook = open_workbook(index_filename)
        sheet = workbook.sheet_by_index(0)
        headers = [cell.value for cell in sheet[0]]
        validate_index_file_parameters(headers, index_filename)
        for row in range(1, sheet.nrows):
            index_row_dict = {headers[i]: sheet.cell_value(row, i) for i in range(len(headers))}
            if validate_index_data(index_row_dict, row):
                data_list.extend(
                    parse_data_file(f'{base_dir}{os.sep}{index_row_dict[title_file_path]}', index_row_dict))

    # 如果既不是xls后缀名，也不是xlsx后缀名，说明输入的索引文件有问题。
    else:
        print("输入的文件不是一个标准的excel，请使用.xls或者.xlsx文件")


def write_to_output_file():
    workbook_total = Workbook()
    sheet_total = workbook_total.active
    for id_r, item in enumerate(data_list):
        for id_c, value in enumerate(item):
            sheet_total.cell(id_r + 1, id_c + 1, value)
            print(value, end='\t')
        print()
    workbook_total.save(output_filename)
    if len(data_list) > 1:
        print(f'数据全部处理完毕，写入文件： {output_filename}')
        workbook_total.close()
    print(f'共写入数据{len(data_list) - 1}条！')


# 定义python的入口文件，代码从此处开始运行⬇️
def main():
    # 获得索引文件
    global base_dir
    global output_filename
    global error_filename
    default_index_filename = f'{base_dir}{os.sep}index.xls'
    print("欢迎使用数据迁移程序！")
    # 引导用户输入索引文件的全路径，并且将值存在index_filename中
    index_filename = input(f'请输入索引文件的路径，按回车键确认(默认路径为：{default_index_filename}):\n')
    # 如果用户什么都没有输入，直接敲了回车，那么就用默认的文件路径
    if index_filename == '':
        index_filename = default_index_filename
    base_dir = f'{os.path.dirname(index_filename)}{os.sep}'
    output_filename = f'{base_dir}{os.sep}total.xlsx'
    output_filename_count = 1
    while os.path.exists(output_filename):
        output_filename = f'{base_dir}{os.sep}total({output_filename_count}).xlsx'
        output_filename_count += 1
    error_filename = f'{base_dir}{os.sep}错误日志.txt'
    input(f'请将所有的数据文件，放置在"{base_dir}"文件夹下，按回车键确认\n')
    # 只有当输入的索引文件真实存在，才会进入到解析逻辑，否则直接报错并且程序结束
    if os.path.exists(index_filename):
        # 解析索引文件
        parse_index_file(index_filename)
        write_to_output_file()
        output_error_list()
    else:
        print(f'找不到"{index_filename}"，请检查文件是否存在！')


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(e)
    finally:
        input("按下任意键结束程序...")
